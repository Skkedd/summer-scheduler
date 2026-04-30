from __future__ import annotations

from datetime import timedelta, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from calendar_math import count_workdays, parse_date_string
from models import ProgressEntry, Room, School, ScheduleSettings, StaffingDay


SINGLE_WORKBOOK_FILE = "Summer Scheduler Workbook.xlsx"

REQUIRED_SHEETS = [
    "Sites",
    "Rooms",
    "Setup",
    "Run Settings",
    "Room Scope",
    "Staffing",
]

REQUIRED_COLUMNS: Dict[str, List[str]] = {
    "Sites": ["Site Name", "Site Order"],
    "Rooms": [
        "Site Name",
        "Building Name",
        "Zone Name",
        "Room Name",
        "Room Order",
        "Total Room SqFt",
        "Carpet SqFt",
        "Carpet Fraction",
        "Strip/Wax Tile SqFt",
        "Strip/Wax Tile Fraction",
        "Scrub-Only Hard Floor SqFt",
        "Scrub-Only Hard Floor Fraction",
    ],
    "Setup": ["Setting", "Value"],
    "Run Settings": ["Setting", "Value"],
    "Room Scope": [
        "Site Name",
        "Building Name",
        "Zone Name",
        "Room Name",
        "Available Day",
        "Include Room",
        "Include Deep Clean",
        "Include Strip",
        "Include Wax",
        "Do Carpet Cleaning",
        "Do Exterior Cleaning",
    ],
    "Staffing": [
        "Day",
        "Available Staff",
        "Carpet Staff Reserved",
        "Absences",
        "Temporary Help",
    ],
    "Progress": [
        "Site Name",
        "Building Name",
        "Zone Name",
        "Room Name",
        "Task",
        "Hours Completed",
    ],
    "Holidays": [
        "Date",
    ],
}

REQUIRED_SETUP_KEYS = [
    "scheduled_shift_hours_per_day",
    "lunch_hours_per_day",
    "break_hours_per_day",
    "setup_hours_per_day",
    "cleanup_hours_per_day",
    "productive_hours_per_staff_per_day",
    "deep_clean_rate_sqft_per_hour",
    "restroom_deep_clean_rate_sqft_per_hour",
    "strip_rate_sqft_per_hour",
    "wax_rate_sqft_per_hour",
    "carpet_rate_sqft_per_hour",
    "exterior_rate_sqft_per_hour",
    "wax_coats",
]

REQUIRED_RUN_SETTING_KEYS = [
    "schedule_name",
    "schedule_start_date",
    "target_end_date",
    "paid_holidays_in_range",
    "work_on_weekends",
    "current_day",
    "include_deep_clean",
    "include_strip",
    "include_wax",
    "include_carpet",
    "include_exterior",
]


def parse_bool(value) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def parse_bool_default_true(value) -> bool:
    """Parse user-facing include toggles where blank means included/on.

    This keeps Room Scope clean: users only need to type FALSE for exceptions.
    """
    if value is None or str(value).strip() == "":
        return True
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def parse_float(value, default: float = 0.0) -> float:
    if value is None or str(value).strip() == "":
        return default
    return float(str(value).strip())


def parse_int(value, default: int = 0) -> int:
    if value is None or str(value).strip() == "":
        return default
    return int(float(str(value).strip()))


def parse_str(value, default: str = "") -> str:
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip()


def get_row_value(row: Dict[str, str], primary: str, *fallbacks: str):
    """Return a row value by preferred column name with optional legacy fallbacks."""
    if primary in row:
        return row.get(primary)
    for fallback in fallbacks:
        if fallback in row:
            return row.get(fallback)
    return None


def _normalize_header(value) -> str:
    return str(value).strip() if value is not None else ""


def _resolve_workbook_path(workbook_path: str | None) -> Path:
    if not workbook_path:
        raise ValueError("No workbook selected.")

    path = Path(workbook_path).expanduser()

    if path.exists() and path.is_dir():
        raise ValueError(
            "The scheduler now uses one workbook file. Select Summer Scheduler Workbook.xlsx, not a folder."
        )

    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Selected workbook must be an Excel .xlsx or .xlsm file.")

    return path.resolve()


def _open_required_workbook(workbook_path: str | None):
    path = _resolve_workbook_path(workbook_path)

    if not path.exists():
        raise ValueError(f"Workbook not found: {path}")

    return load_workbook(path, data_only=True)


def _find_header_row(ws, required_headers: Optional[List[str]] = None) -> Tuple[int, List[str]]:
    required_headers = required_headers or []

    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        headers = [_normalize_header(cell) for cell in row]
        header_set = {h for h in headers if h}

        if required_headers:
            if all(header in header_set for header in required_headers):
                return row_index, headers
        elif any(headers):
            return row_index, headers

    required_text = ", ".join(required_headers) if required_headers else "any headers"
    raise ValueError(
        f"Could not find header row in sheet '{ws.title}' requiring: {required_text}"
    )


def _sheet_to_dict_rows(ws, required_headers: Optional[List[str]] = None) -> List[Dict[str, str]]:
    header_row, headers = _find_header_row(ws, required_headers)
    data_rows: List[Dict[str, str]] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        record: Dict[str, str] = {}
        has_any_value = False

        for i, header in enumerate(headers):
            if not header:
                continue

            value = row[i] if i < len(row) else ""
            if value is not None and str(value).strip() != "":
                has_any_value = True

            record[header] = "" if value is None else str(value).strip()

        if has_any_value:
            data_rows.append(record)

    return data_rows


def _load_key_value_sheet(wb, sheet_name: str) -> Dict[str, str]:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook is missing required sheet: {sheet_name}")

    rows = _sheet_to_dict_rows(wb[sheet_name], REQUIRED_COLUMNS[sheet_name])
    raw: Dict[str, str] = {}

    for row in rows:
        key = parse_str(row.get("Setting"))
        value = parse_str(row.get("Value"))
        if key:
            raw[key] = value

    return raw


def observed_holiday_date(holiday_date: date) -> date:
    """Return the normal observed weekday for a holiday.

    Saturday holidays are observed Friday. Sunday holidays are observed Monday.
    Weekday holidays are observed on their actual date.
    """
    if holiday_date.weekday() == 5:  # Saturday
        return holiday_date - timedelta(days=1)
    if holiday_date.weekday() == 6:  # Sunday
        return holiday_date + timedelta(days=1)
    return holiday_date


def load_holidays(workbook_path: str | None = None) -> set[date]:
    """Load actual non-work dates from the optional Holidays sheet.

    Expected columns:
    - Date: actual holiday date
    Optional columns:
    - Observed Date: override observed/closure date
    - Counts As Non-Workday: blank/TRUE means skip this date; FALSE ignores it
    - Holiday Name or Name: label only

    If Observed Date is blank, weekend holidays are automatically observed:
    Saturday -> Friday, Sunday -> Monday.
    """
    wb = _open_required_workbook(workbook_path)

    if "Holidays" not in wb.sheetnames:
        return set()

    rows = _sheet_to_dict_rows(wb["Holidays"], ["Date"])
    holidays: set[date] = set()

    for row in rows:
        if not parse_bool_default_true(row.get("Counts As Non-Workday")):
            continue

        raw_date = parse_str(row.get("Date"))
        raw_observed = parse_str(row.get("Observed Date"))

        if not raw_date and not raw_observed:
            continue

        try:
            if raw_observed:
                holidays.add(parse_date_string(raw_observed))
            else:
                holidays.add(observed_holiday_date(parse_date_string(raw_date)))
        except Exception:
            # Validation reports bad date values. Loading skips bad holiday rows so the app
            # can still show all validation errors cleanly.
            continue

    return holidays


def _calculate_workdays(
    start_date_str: str,
    end_date_str: str,
    include_weekends: bool,
    holidays: set[date],
    legacy_paid_holidays: int = 0,
) -> int:
    workdays = count_workdays(
        start_date_str,
        end_date_str,
        work_on_weekends=include_weekends,
        holidays=holidays,
    )

    # Backward compatibility for old workbooks that do not have a Holidays sheet yet.
    # Once the Holidays sheet is present, exact dates become the source of truth.
    if not holidays and legacy_paid_holidays > 0:
        workdays -= legacy_paid_holidays

    return max(workdays, 0)


def load_settings(workbook_path: str | None = None) -> ScheduleSettings:
    wb = _open_required_workbook(workbook_path)

    assumptions = _load_key_value_sheet(wb, "Setup")
    run_input = _load_key_value_sheet(wb, "Run Settings")

    scheduled_shift_hours_per_day = parse_float(
        assumptions.get("scheduled_shift_hours_per_day"),
        8.5,
    )
    lunch_hours_per_day = parse_float(assumptions.get("lunch_hours_per_day"), 0.5)
    break_hours_per_day = parse_float(assumptions.get("break_hours_per_day"), 0.5)
    setup_hours_per_day = parse_float(assumptions.get("setup_hours_per_day"), 0.25)
    cleanup_hours_per_day = parse_float(assumptions.get("cleanup_hours_per_day"), 0.5)

    calculated_productive_default = (
        scheduled_shift_hours_per_day
        - lunch_hours_per_day
        - break_hours_per_day
        - setup_hours_per_day
        - cleanup_hours_per_day
    )
    if calculated_productive_default < 0:
        calculated_productive_default = 0.0

    schedule_start_date = parse_str(run_input.get("schedule_start_date"), "2026-06-01")
    target_end_date = parse_str(run_input.get("target_end_date"), "")
    work_on_weekends = parse_bool(run_input.get("work_on_weekends", "False"))
    paid_holidays = parse_int(run_input.get("paid_holidays_in_range"), 0)
    holidays = load_holidays(workbook_path)

    explicit_target_end_day = parse_int(run_input.get("target_end_day"), 0)
    if explicit_target_end_day > 0:
        target_end_day = explicit_target_end_day
    elif target_end_date:
        target_end_day = _calculate_workdays(
            schedule_start_date,
            target_end_date,
            work_on_weekends,
            holidays,
            legacy_paid_holidays=paid_holidays,
        )
    else:
        target_end_day = 20

    settings = ScheduleSettings(
        schedule_name=parse_str(run_input.get("schedule_name"), "Summer Schedule"),
        target_end_day=target_end_day,
        scheduled_shift_hours_per_day=scheduled_shift_hours_per_day,
        lunch_hours_per_day=lunch_hours_per_day,
        break_hours_per_day=break_hours_per_day,
        setup_hours_per_day=setup_hours_per_day,
        cleanup_hours_per_day=cleanup_hours_per_day,
        productive_hours_per_staff_per_day=parse_float(
            assumptions.get("productive_hours_per_staff_per_day"),
            calculated_productive_default,
        ),
        current_day=parse_int(run_input.get("current_day"), 1),
        include_deep_clean=parse_bool(run_input.get("include_deep_clean", "True")),
        include_strip=parse_bool(run_input.get("include_strip", "True")),
        include_wax=parse_bool(run_input.get("include_wax", "True")),
        include_carpet=parse_bool(run_input.get("include_carpet", "True")),
        include_exterior=parse_bool(run_input.get("include_exterior", "False")),
        deep_clean_rate_sqft_per_hour=parse_float(
            assumptions.get("deep_clean_rate_sqft_per_hour"),
            400.0,
        ),
        restroom_deep_clean_rate_sqft_per_hour=parse_float(
            assumptions.get("restroom_deep_clean_rate_sqft_per_hour"),
            250.0,
        ),
        strip_rate_sqft_per_hour=parse_float(
            assumptions.get("strip_rate_sqft_per_hour"),
            300.0,
        ),
        wax_rate_sqft_per_hour=parse_float(
            assumptions.get("wax_rate_sqft_per_hour"),
            600.0,
        ),
        carpet_rate_sqft_per_hour=parse_float(
            assumptions.get("carpet_rate_sqft_per_hour"),
            500.0,
        ),
        exterior_rate_sqft_per_hour=parse_float(
            assumptions.get("exterior_rate_sqft_per_hour"),
            1000.0,
        ),
        wax_coats=parse_int(assumptions.get("wax_coats"), 3),
        transition_hours_per_school=parse_float(
            assumptions.get("transition_hours_per_school"),
            0.0,
        ),
        day_start_time=parse_str(assumptions.get("day_start_time"), "7:30 AM"),
        work_start_time=parse_str(assumptions.get("work_start_time"), "7:45 AM"),
        first_break_time=parse_str(assumptions.get("first_break_time"), "10:00 AM"),
        lunch_time=parse_str(assumptions.get("lunch_time"), "12:00 PM"),
        second_break_time=parse_str(assumptions.get("second_break_time"), "2:00 PM"),
        cleanup_start_time=parse_str(assumptions.get("cleanup_start_time"), "3:30 PM"),
        day_end_time=parse_str(assumptions.get("day_end_time"), "4:00 PM"),
        schedule_start_date=schedule_start_date,
        target_end_date=target_end_date,
        work_on_weekends=work_on_weekends,
        paid_holidays_in_range=paid_holidays,
    )

    settings.validate_or_normalize()
    return settings


def _resolve_floor_sqft(
    total_room_sqft: float,
    exact_sqft: float,
    fraction: float,
) -> float:
    if exact_sqft > 0:
        return exact_sqft
    if fraction > 0 and total_room_sqft > 0:
        return total_room_sqft * fraction
    return 0.0


def load_rooms(workbook_path: str | None = None) -> Tuple[List[Room], List[School]]:
    wb = _open_required_workbook(workbook_path)

    for sheet in ["Sites", "Rooms", "Room Scope"]:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {sheet}")

    site_rows = _sheet_to_dict_rows(wb["Sites"], REQUIRED_COLUMNS["Sites"])
    room_rows = _sheet_to_dict_rows(wb["Rooms"], REQUIRED_COLUMNS["Rooms"])
    scope_rows = _sheet_to_dict_rows(wb["Room Scope"], REQUIRED_COLUMNS["Room Scope"])

    site_order_map: Dict[str, int] = {}
    for row in site_rows:
        site_name = parse_str(row.get("Site Name"))
        if site_name:
            site_order_map[site_name] = parse_int(row.get("Site Order"), 999)

    scope_map: Dict[Tuple[str, str, str, str], Dict[str, str]] = {}
    for row in scope_rows:
        key = (
            parse_str(row.get("Site Name")),
            parse_str(row.get("Building Name")),
            parse_str(row.get("Zone Name")),
            parse_str(row.get("Room Name")),
        )
        scope_map[key] = row

    rooms: List[Room] = []
    schools_by_name: Dict[str, School] = {}

    for row in room_rows:
        school_name = parse_str(row.get("Site Name"))
        building_name = parse_str(row.get("Building Name"))
        zone_name = parse_str(row.get("Zone Name"))
        room_name = parse_str(row.get("Room Name"))
        room_order = parse_int(row.get("Room Order"), 999)
        total_room_sqft = parse_float(row.get("Total Room SqFt"), 0.0)

        carpet_sqft = _resolve_floor_sqft(
            total_room_sqft=total_room_sqft,
            exact_sqft=parse_float(row.get("Carpet SqFt"), 0.0),
            fraction=parse_float(row.get("Carpet Fraction"), 0.0),
        )
        tile_strip_wax_sqft = _resolve_floor_sqft(
            total_room_sqft=total_room_sqft,
            exact_sqft=parse_float(row.get("Strip/Wax Tile SqFt"), 0.0),
            fraction=parse_float(row.get("Strip/Wax Tile Fraction"), 0.0),
        )
        scrub_only_hard_floor_sqft = _resolve_floor_sqft(
            total_room_sqft=total_room_sqft,
            exact_sqft=parse_float(row.get("Scrub-Only Hard Floor SqFt"), 0.0),
            fraction=parse_float(row.get("Scrub-Only Hard Floor Fraction"), 0.0),
        )

        key = (school_name, building_name, zone_name, room_name)
        scope = scope_map.get(key, {})

        if not parse_bool_default_true(scope.get("Include Room")):
            continue

        room = Room(
            school_name=school_name,
            school_order=site_order_map.get(school_name, 999),
            building_name=building_name,
            zone_name=zone_name,
            room_name=room_name,
            room_order=room_order,
            total_room_sqft=total_room_sqft,
            carpet_sqft=carpet_sqft,
            tile_strip_wax_sqft=tile_strip_wax_sqft,
            scrub_only_hard_floor_sqft=scrub_only_hard_floor_sqft,
            room_use=parse_str(row.get("Room Use"), ""),
            available_day=parse_int(scope.get("Available Day"), 1),
            include_deep_clean=parse_bool_default_true(scope.get("Include Deep Clean")),
            include_strip=parse_bool_default_true(scope.get("Include Strip")),
            include_wax=parse_bool_default_true(scope.get("Include Wax")),
            include_carpet=parse_bool_default_true(
                get_row_value(scope, "Do Carpet Cleaning", "Include Carpet")
            ),
            include_exterior=parse_bool_default_true(
                get_row_value(scope, "Do Exterior Cleaning", "Include Exterior")
            ),
            notes=parse_str(scope.get("Notes"), ""),
        )
        rooms.append(room)

        if room.school_name not in schools_by_name:
            schools_by_name[room.school_name] = School(
                name=room.school_name,
                order=room.school_order,
            )

        schools_by_name[room.school_name].add_room(room)

    schools = sorted(schools_by_name.values(), key=lambda s: (s.order, s.name))
    return rooms, schools


def load_staffing(workbook_path: str | None = None) -> List[StaffingDay]:
    wb = _open_required_workbook(workbook_path)

    if "Staffing" not in wb.sheetnames:
        raise ValueError("Workbook is missing required sheet: Staffing")

    rows = _sheet_to_dict_rows(wb["Staffing"], REQUIRED_COLUMNS["Staffing"])
    staffing_days: List[StaffingDay] = []

    for row in rows:
        staffing_days.append(
            StaffingDay(
                day=parse_int(row.get("Day"), 1),
                available_staff=parse_int(row.get("Available Staff"), 0),
                carpet_staff_reserved=parse_int(row.get("Carpet Staff Reserved"), 0),
                absences=parse_int(row.get("Absences"), 0),
                temporary_help=parse_int(row.get("Temporary Help"), 0),
            )
        )

    staffing_days.sort(key=lambda s: s.day)
    return staffing_days


def load_progress(workbook_path: str | None = None) -> List[ProgressEntry]:
    wb = _open_required_workbook(workbook_path)

    if "Progress" not in wb.sheetnames:
        return []

    rows = _sheet_to_dict_rows(wb["Progress"], REQUIRED_COLUMNS["Progress"])
    progress_entries: List[ProgressEntry] = []

    for row in rows:
        school_name = parse_str(row.get("Site Name"))
        room_name = parse_str(row.get("Room Name"))
        phase_name = parse_str(row.get("Task"))
        hours_completed = parse_str(row.get("Hours Completed"))

        if not school_name or not room_name or not phase_name or not hours_completed:
            continue

        progress_entries.append(
            ProgressEntry(
                school_name=school_name,
                building_name=parse_str(row.get("Building Name"), ""),
                zone_name=parse_str(row.get("Zone Name"), ""),
                room_name=room_name,
                phase_name=phase_name,
                hours_completed=float(hours_completed),
            )
        )

    return progress_entries


def _append_missing_columns(errors: List[str], wb, sheet_name: str) -> None:
    if sheet_name not in wb.sheetnames:
        errors.append(f"Missing required sheet: {sheet_name}")
        return

    try:
        _find_header_row(wb[sheet_name], REQUIRED_COLUMNS[sheet_name])
    except Exception as exc:
        errors.append(f"{sheet_name}: {exc}")


def _check_required_key(errors: List[str], raw: Dict[str, str], sheet_name: str, key: str) -> None:
    if key not in raw or parse_str(raw.get(key)) == "":
        errors.append(f"{sheet_name}: Missing required setting '{key}'")


def _check_float(
    errors: List[str],
    sheet_name: str,
    row_label: str,
    field: str,
    value,
    allow_blank: bool = True,
    minimum: Optional[float] = None,
    maximum: Optional[float] = None,
) -> None:
    raw = parse_str(value)
    if raw == "":
        if not allow_blank:
            errors.append(f"{sheet_name} {row_label}: {field} is required")
        return
    try:
        number = float(raw)
    except Exception:
        errors.append(f"{sheet_name} {row_label}: {field} must be a number")
        return
    if minimum is not None and number < minimum:
        errors.append(f"{sheet_name} {row_label}: {field} must be at least {minimum}")
    if maximum is not None and number > maximum:
        errors.append(f"{sheet_name} {row_label}: {field} must be no more than {maximum}")


def _check_int(
    errors: List[str],
    sheet_name: str,
    row_label: str,
    field: str,
    value,
    allow_blank: bool = True,
    minimum: Optional[int] = None,
) -> None:
    raw = parse_str(value)
    if raw == "":
        if not allow_blank:
            errors.append(f"{sheet_name} {row_label}: {field} is required")
        return
    try:
        number = int(float(raw))
    except Exception:
        errors.append(f"{sheet_name} {row_label}: {field} must be a whole number")
        return
    if minimum is not None and number < minimum:
        errors.append(f"{sheet_name} {row_label}: {field} must be at least {minimum}")


def _check_bool(
    errors: List[str],
    sheet_name: str,
    row_label: str,
    field: str,
    value,
    allow_blank: bool = True,
) -> None:
    raw = parse_str(value)
    if raw == "":
        if not allow_blank:
            errors.append(f"{sheet_name} {row_label}: {field} is required")
        return
    if raw.strip().lower() not in {"true", "false", "1", "0", "yes", "no", "y", "n"}:
        errors.append(f"{sheet_name} {row_label}: {field} must be True/False")


def _validate_settings(errors: List[str], wb) -> None:
    try:
        setup = _load_key_value_sheet(wb, "Setup")
        run_settings = _load_key_value_sheet(wb, "Run Settings")
    except Exception as exc:
        errors.append(str(exc))
        return

    for key in REQUIRED_SETUP_KEYS:
        _check_required_key(errors, setup, "Setup", key)
    for key in REQUIRED_RUN_SETTING_KEYS:
        _check_required_key(errors, run_settings, "Run Settings", key)

    for key in [
        "scheduled_shift_hours_per_day",
        "lunch_hours_per_day",
        "break_hours_per_day",
        "setup_hours_per_day",
        "cleanup_hours_per_day",
        "productive_hours_per_staff_per_day",
        "deep_clean_rate_sqft_per_hour",
        "strip_rate_sqft_per_hour",
        "wax_rate_sqft_per_hour",
        "carpet_rate_sqft_per_hour",
        "exterior_rate_sqft_per_hour",
        "restroom_deep_clean_rate_sqft_per_hour",
        "transition_hours_per_school",
    ]:
        _check_float(errors, "Setup", f"setting '{key}'", "Value", setup.get(key), allow_blank=False, minimum=0)

    _check_int(errors, "Setup", "setting 'wax_coats'", "Value", setup.get("wax_coats"), allow_blank=False, minimum=1)
    _check_int(errors, "Run Settings", "setting 'current_day'", "Value", run_settings.get("current_day"), allow_blank=False, minimum=1)
    _check_int(errors, "Run Settings", "setting 'paid_holidays_in_range'", "Value", run_settings.get("paid_holidays_in_range"), allow_blank=False, minimum=0)

    for key in [
        "work_on_weekends",
        "include_deep_clean",
        "include_strip",
        "include_wax",
        "include_carpet",
        "include_exterior",
    ]:
        _check_bool(errors, "Run Settings", f"setting '{key}'", "Value", run_settings.get(key), allow_blank=False)

    try:
        parse_date_string(parse_str(run_settings.get("schedule_start_date")))
    except Exception as exc:
        errors.append(f"Run Settings: schedule_start_date is invalid: {exc}")

    target_end_date = parse_str(run_settings.get("target_end_date"))
    if target_end_date:
        try:
            parse_date_string(target_end_date)
        except Exception as exc:
            errors.append(f"Run Settings: target_end_date is invalid: {exc}")



def _validate_holidays(errors: List[str], wb) -> None:
    if "Holidays" not in wb.sheetnames:
        return

    try:
        rows = _sheet_to_dict_rows(wb["Holidays"], ["Date"])
    except Exception as exc:
        errors.append(f"Holidays: {exc}")
        return

    seen_observed_dates = set()
    for index, row in enumerate(rows, start=2):
        label = f"row {index}"
        raw_date = parse_str(row.get("Date"))
        raw_observed = parse_str(row.get("Observed Date"))

        if not raw_date and not raw_observed:
            continue

        if row.get("Counts As Non-Workday") is not None:
            _check_bool(
                errors,
                "Holidays",
                label,
                "Counts As Non-Workday",
                row.get("Counts As Non-Workday"),
                allow_blank=True,
            )

        try:
            actual_date = parse_date_string(raw_date) if raw_date else None
        except Exception as exc:
            errors.append(f"Holidays {label}: Date is invalid: {exc}")
            actual_date = None

        try:
            if raw_observed:
                observed = parse_date_string(raw_observed)
            elif actual_date:
                observed = observed_holiday_date(actual_date)
            else:
                observed = None
        except Exception as exc:
            errors.append(f"Holidays {label}: Observed Date is invalid: {exc}")
            observed = None

        if observed:
            if observed in seen_observed_dates:
                errors.append(f"Holidays {label}: Duplicate observed non-workday {observed.isoformat()}")
            seen_observed_dates.add(observed)

def _validate_rows(errors: List[str], wb) -> None:
    try:
        site_rows = _sheet_to_dict_rows(wb["Sites"], REQUIRED_COLUMNS["Sites"])
        room_rows = _sheet_to_dict_rows(wb["Rooms"], REQUIRED_COLUMNS["Rooms"])
        scope_rows = _sheet_to_dict_rows(wb["Room Scope"], REQUIRED_COLUMNS["Room Scope"])
        staffing_rows = _sheet_to_dict_rows(wb["Staffing"], REQUIRED_COLUMNS["Staffing"])
    except Exception as exc:
        errors.append(str(exc))
        return

    site_names = set()
    for index, row in enumerate(site_rows, start=2):
        label = f"row {index}"
        site_name = parse_str(row.get("Site Name"))
        if not site_name:
            errors.append(f"Sites {label}: Site Name is required")
        else:
            site_names.add(site_name)
        _check_int(errors, "Sites", label, "Site Order", row.get("Site Order"), allow_blank=False, minimum=1)

    room_keys = set()
    for index, row in enumerate(room_rows, start=2):
        label = f"row {index}"
        site_name = parse_str(row.get("Site Name"))
        building = parse_str(row.get("Building Name"))
        zone = parse_str(row.get("Zone Name"))
        room = parse_str(row.get("Room Name"))

        for field in ["Site Name", "Building Name", "Zone Name", "Room Name"]:
            if not parse_str(row.get(field)):
                errors.append(f"Rooms {label}: {field} is required")

        if site_name and site_names and site_name not in site_names:
            errors.append(f"Rooms {label}: Site Name '{site_name}' does not exist on Sites sheet")

        key = (site_name, building, zone, room)
        if all(key):
            if key in room_keys:
                errors.append(f"Rooms {label}: Duplicate room key {key}")
            room_keys.add(key)

        _check_int(errors, "Rooms", label, "Room Order", row.get("Room Order"), allow_blank=False, minimum=1)
        _check_float(errors, "Rooms", label, "Total Room SqFt", row.get("Total Room SqFt"), allow_blank=False, minimum=0)
        for field in ["Carpet SqFt", "Strip/Wax Tile SqFt", "Scrub-Only Hard Floor SqFt"]:
            _check_float(errors, "Rooms", label, field, row.get(field), allow_blank=True, minimum=0)
        for field in ["Carpet Fraction", "Strip/Wax Tile Fraction", "Scrub-Only Hard Floor Fraction"]:
            _check_float(errors, "Rooms", label, field, row.get(field), allow_blank=True, minimum=0, maximum=1)

    for index, row in enumerate(scope_rows, start=2):
        label = f"row {index}"
        key = (
            parse_str(row.get("Site Name")),
            parse_str(row.get("Building Name")),
            parse_str(row.get("Zone Name")),
            parse_str(row.get("Room Name")),
        )
        if all(key) and room_keys and key not in room_keys:
            errors.append(f"Room Scope {label}: Room {key} does not exist on Rooms sheet")
        _check_int(errors, "Room Scope", label, "Available Day", row.get("Available Day"), allow_blank=True, minimum=1)
        for field in [
            "Include Room",
            "Include Deep Clean",
            "Include Strip",
            "Include Wax",
            "Do Carpet Cleaning",
            "Do Exterior Cleaning",
        ]:
            _check_bool(errors, "Room Scope", label, field, row.get(field), allow_blank=True)

    seen_staffing_days = set()
    for index, row in enumerate(staffing_rows, start=2):
        label = f"row {index}"
        day_raw = row.get("Day")
        _check_int(errors, "Staffing", label, "Day", day_raw, allow_blank=False, minimum=1)
        day = parse_int(day_raw, 0)
        if day > 0:
            if day in seen_staffing_days:
                errors.append(f"Staffing {label}: Duplicate Day {day}")
            seen_staffing_days.add(day)

        for field in ["Available Staff", "Carpet Staff Reserved", "Absences", "Temporary Help"]:
            _check_int(errors, "Staffing", label, field, row.get(field), allow_blank=False, minimum=0)

    if "Progress" in wb.sheetnames:
        try:
            progress_rows = _sheet_to_dict_rows(wb["Progress"], REQUIRED_COLUMNS["Progress"])
            for index, row in enumerate(progress_rows, start=2):
                label = f"row {index}"
                if parse_str(row.get("Hours Completed")):
                    _check_float(
                        errors,
                        "Progress",
                        label,
                        "Hours Completed",
                        row.get("Hours Completed"),
                        allow_blank=True,
                        minimum=0,
                    )
        except Exception as exc:
            errors.append(str(exc))


def validate_workbook(workbook_path: str | None = None) -> List[str]:
    errors: List[str] = []

    try:
        path = _resolve_workbook_path(workbook_path)
    except Exception as exc:
        return [str(exc)]

    if not path.exists():
        return [f"Missing required workbook: {path.name}"]

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        return [f"Could not open workbook: {exc}"]

    for sheet in REQUIRED_SHEETS:
        if sheet not in wb.sheetnames:
            errors.append(f"Missing required sheet: {sheet}")

    if errors:
        return errors

    sheet_names = ["Sites", "Rooms", "Setup", "Run Settings", "Room Scope", "Staffing"]
    if "Progress" in wb.sheetnames:
        sheet_names.append("Progress")
    if "Holidays" in wb.sheetnames:
        sheet_names.append("Holidays")

    for sheet_name in sheet_names:
        _append_missing_columns(errors, wb, sheet_name)

    if errors:
        return errors

    _validate_settings(errors, wb)
    _validate_holidays(errors, wb)
    _validate_rows(errors, wb)

    return errors
