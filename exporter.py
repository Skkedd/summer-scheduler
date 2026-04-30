from __future__ import annotations

import os
from pathlib import Path
from datetime import date
from typing import Optional, Sequence, Set

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from calendar_math import format_date_label, workday_to_date
from models import ScheduleResult, ScheduleSettings


DISTRICT_FILE = "District Facility Data.xlsx"
ASSUMPTIONS_FILE = "Cleaning Planning Assumptions.xlsx"
RUN_INPUT_FILE = "Summer Scheduler Run Input.xlsx"
FULL_RESULTS_FILE = "Full Schedule Results.xlsx"
WORKER_EXPORT_FILE = "Worker Schedule Export.xlsx"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
INSTRUCTION_FILL = PatternFill(fill_type="solid", fgColor="FFF4CC")


def ensure_output_folder(folder_path: str = "output") -> None:
    os.makedirs(folder_path, exist_ok=True)


def _bold_header(ws, row_num: int = 1) -> None:
    for cell in ws[row_num]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_instruction_row(ws, row_num: int = 1) -> None:
    for cell in ws[row_num]:
        cell.font = Font(bold=True)
        cell.fill = INSTRUCTION_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _freeze_header(ws, cell: str = "A2") -> None:
    ws.freeze_panes = cell


def _set_column_widths(ws, widths: dict[str, float]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _autosize_from_headers(ws, min_width: int = 14, max_width: int = 34) -> None:
    for col_idx, cell in enumerate(ws[1], start=1):
        value = str(cell.value or "")
        width = min(max(len(value) + 4, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _selected_site_set(selected_sites: Optional[Sequence[str]]) -> Optional[Set[str]]:
    if not selected_sites:
        return None
    return {item.strip() for item in selected_sites if str(item).strip()}


def _task_matches_sites(task, selected_sites: Optional[Set[str]]) -> bool:
    if selected_sites is None:
        return True
    return task.school_name in selected_sites


def _worklog_matches_sites(item, selected_sites: Optional[Set[str]]) -> bool:
    if selected_sites is None:
        return True
    return item.school_name in selected_sites


def _filtered_days(result: ScheduleResult, selected_sites: Optional[Set[str]]):
    filtered = []
    for day in result.days:
        matching_log = [item for item in day.work_log if _worklog_matches_sites(item, selected_sites)]
        if matching_log:
            filtered.append((day, matching_log))
    return filtered


def _filtered_tasks(result: ScheduleResult, selected_sites: Optional[Set[str]]):
    return [task for task in result.task_items if _task_matches_sites(task, selected_sites)]


def export_result_workbook(
    result: ScheduleResult,
    settings: ScheduleSettings,
    folder_path: str = "output",
    selected_sites: Optional[Sequence[str]] = None,
    holidays: Optional[Set[date]] = None,
) -> str:
    """
    Export the full schedule workbook.

    This accepts either an output folder or a full .xlsx/.xlsm file path.
    The UI uses a Save As dialog and passes a full file path.
    """
    target = Path(folder_path)

    if target.suffix.lower() in {".xlsx", ".xlsm"}:
        target.parent.mkdir(parents=True, exist_ok=True)
        file_path = str(target)
    else:
        ensure_output_folder(str(target))
        file_path = os.path.join(str(target), FULL_RESULTS_FILE)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    selected_site_set = _selected_site_set(selected_sites)
    filtered_tasks = _filtered_tasks(result, selected_site_set)
    filtered_days = _filtered_days(result, selected_site_set)
    holidays = holidays or set()

    ws = wb.create_sheet("Summary")
    ws.append(["Field", "Value"])
    _bold_header(ws)

    filtered_total_planned = round(sum(task.total_hours for task in filtered_tasks), 2)
    filtered_remaining = round(sum(task.remaining_hours for task in filtered_tasks), 2)
    filtered_used = round(sum(item.hours_done for _, log in filtered_days for item in log), 2)

    summary_rows = [
        ("Schedule Name", result.schedule_name),
        ("Schedule Start Date", settings.schedule_start_date),
        ("Target End Date", settings.target_end_date),
        ("Work On Weekends", settings.work_on_weekends),
        ("Legacy Paid Holidays In Range", settings.paid_holidays_in_range),
        ("Holiday Dates Used", ", ".join(sorted(d.isoformat() for d in holidays)) if holidays else "None"),
        ("Target End Day", result.target_end_day),
        ("Current Day", result.current_day),
        ("Projected Finish Day", result.finish_day),
        (
            "Projected Finish Date",
            format_date_label(
                workday_to_date(
                    settings.schedule_start_date,
                    result.finish_day,
                    settings.work_on_weekends,
                    holidays=holidays,
                )
            ),
        ),
        ("Deadline Met", result.met_deadline),
        ("Export Scope", ", ".join(sorted(selected_site_set)) if selected_site_set else "All Sites"),
        ("Filtered Total Planned Hours", filtered_total_planned),
        ("Filtered Total Used Hours", filtered_used),
        ("Filtered Remaining Backlog Hours", filtered_remaining),
        ("Recommendation Status", result.recommendation.status_label),
        ("Bottleneck Type", result.recommendation.bottleneck_type),
        ("Recommended Action", result.recommendation.recommended_action),
    ]
    for row in summary_rows:
        ws.append(list(row))
    _set_column_widths(ws, {"A": 30, "B": 50})

    ws = wb.create_sheet("Days")
    ws.append(
        [
            "Workday Number",
            "Work Date",
            "Active Site",
            "Day Note",
            "Effective Staff",
            "Cleaning Staff",
            "Carpet Staff",
            "Cleaning Capacity Hours",
            "Carpet Capacity Hours",
            "Total Capacity Hours",
            "Used Hours",
            "Unused Hours",
            "Matching Work Items",
        ]
    )
    _bold_header(ws)
    _freeze_header(ws)

    for day, matching_log in filtered_days:
        work_date = workday_to_date(
            settings.schedule_start_date,
            day.day,
            settings.work_on_weekends,
            holidays=holidays,
        )
        ws.append(
            [
                day.day,
                format_date_label(work_date),
                day.active_school_name,
                day.status_note,
                day.effective_staff,
                day.general_staff,
                day.carpet_staff,
                round(day.general_capacity, 2),
                round(day.carpet_capacity, 2),
                round(day.daily_capacity, 2),
                round(sum(item.hours_done for item in matching_log), 2),
                round(day.daily_capacity - sum(item.hours_done for item in matching_log), 2),
                len(matching_log),
            ]
        )
    _autosize_from_headers(ws)

    ws = wb.create_sheet("Work Log")
    ws.append(
        [
            "Workday Number",
            "Work Date",
            "Crew Type",
            "Site",
            "Building",
            "Zone",
            "Room",
            "Task",
            "Available Day",
            "Hours Done",
            "Notes",
        ]
    )
    _bold_header(ws)
    _freeze_header(ws)

    for day, matching_log in filtered_days:
        work_date = workday_to_date(
            settings.schedule_start_date,
            day.day,
            settings.work_on_weekends,
            holidays=holidays,
        )
        for item in matching_log:
            ws.append(
                [
                    day.day,
                    format_date_label(work_date),
                    item.crew_type,
                    item.school_name,
                    item.building_name,
                    item.zone_name,
                    item.room_name,
                    item.phase_name,
                    item.available_day if item.available_day is not None else "",
                    round(item.hours_done, 2),
                    item.note,
                ]
            )
    _autosize_from_headers(ws)

    ws = wb.create_sheet("Tasks")
    ws.append(
        [
            "Site",
            "Building",
            "Zone",
            "Room",
            "Task",
            "Available Day",
            "Available Date",
            "Total Hours",
            "Remaining Hours",
            "Status",
            "Notes",
        ]
    )
    _bold_header(ws)
    _freeze_header(ws)

    for task in filtered_tasks:
        if task.remaining_hours <= 0:
            status = "Complete"
        elif task.remaining_hours < task.total_hours:
            status = "In Progress"
        else:
            status = "Not Started"

        available_date = workday_to_date(
            settings.schedule_start_date,
            task.available_day,
            settings.work_on_weekends,
            holidays=holidays,
        )

        ws.append(
            [
                task.school_name,
                task.building_name,
                task.zone_name,
                task.room_name,
                task.phase_name,
                task.available_day,
                format_date_label(available_date),
                round(task.total_hours, 2),
                round(task.remaining_hours, 2),
                status,
                task.notes,
            ]
        )
    _autosize_from_headers(ws)

    if holidays:
        ws = wb.create_sheet("Holidays Used")
        ws.append(["Observed Non-Work Date"])
        _bold_header(ws)
        for holiday in sorted(holidays):
            ws.append([format_date_label(holiday)])
        _autosize_from_headers(ws)

    wb.save(file_path)
    return file_path


def export_worker_schedule_workbook(
    result: ScheduleResult,
    settings: ScheduleSettings,
    folder_path: str = "output",
    selected_sites: Optional[Sequence[str]] = None,
    holidays: Optional[Set[date]] = None,
) -> str:
    ensure_output_folder(folder_path)
    file_path = os.path.join(folder_path, WORKER_EXPORT_FILE)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    selected_site_set = _selected_site_set(selected_sites)
    filtered_days = _filtered_days(result, selected_site_set)
    holidays = holidays or set()

    ws = wb.create_sheet("Daily Assignment Overview")
    ws.append(
        [
            "Date",
            "Workday Number",
            "Site",
            "Crew Type",
            "Main Task",
            "Hours",
            "Notes",
        ]
    )
    _bold_header(ws)
    _freeze_header(ws)

    for day, matching_log in filtered_days:
        work_date = workday_to_date(
            settings.schedule_start_date,
            day.day,
            settings.work_on_weekends,
            holidays=holidays,
        )

        for item in matching_log:
            main_task = f"{item.phase_name} - {item.room_name}"
            ws.append(
                [
                    format_date_label(work_date),
                    day.day,
                    item.school_name,
                    item.crew_type,
                    main_task,
                    round(item.hours_done, 2),
                    item.note,
                ]
            )
    _autosize_from_headers(ws)

    ws = wb.create_sheet("Detailed Daily Work")
    ws.append(
        [
            "Date",
            "Site",
            "Building",
            "Zone",
            "Room",
            "Work Type",
            "Crew Type",
            "Hours",
            "Notes",
        ]
    )
    _bold_header(ws)
    _freeze_header(ws)

    for day, matching_log in filtered_days:
        work_date = workday_to_date(
            settings.schedule_start_date,
            day.day,
            settings.work_on_weekends,
            holidays=holidays,
        )

        for item in matching_log:
            ws.append(
                [
                    format_date_label(work_date),
                    item.school_name,
                    item.building_name,
                    item.zone_name,
                    item.room_name,
                    item.phase_name,
                    item.crew_type,
                    round(item.hours_done, 2),
                    item.note,
                ]
            )
    _autosize_from_headers(ws)

    ws = wb.create_sheet("Summary")
    ws.append(["Field", "Value"])
    _bold_header(ws)
    _set_column_widths(ws, {"A": 28, "B": 50})

    ws.append(["Schedule Name", result.schedule_name])
    ws.append(["Export Scope", ", ".join(sorted(selected_site_set)) if selected_site_set else "All Sites"])
    ws.append(["Projected Finish Day", result.finish_day])
    ws.append(
        [
            "Projected Finish Date",
            format_date_label(
                workday_to_date(
                    settings.schedule_start_date,
                    result.finish_day,
                    settings.work_on_weekends,
                    holidays=holidays,
                )
            ),
        ]
    )
    ws.append(["Deadline Met", result.met_deadline])
    ws.append(["Recommendation Status", result.recommendation.status_label])
    ws.append(["Recommended Action", result.recommendation.recommended_action])

    wb.save(file_path)
    return file_path




CLEANING_METHOD_COLUMNS = [
    "Method Name",
    "Room Use",
    "Step Order",
    "Step Title",
    "Task Detail",
    "Estimated Minutes",
    "Staff Needed",
    "Team/Role",
    "Notes",
    "Print on Step Sheet",
]

DEFAULT_CLEANING_METHOD_ROWS = [
    [
        "Classroom Deep Clean", "Classroom", 1, "High-to-Low Dusting",
        "Dust vents, light fixtures, projector/speakers, tops of boards, blinds, cabinets, ledges, trim and baseboards.",
        10, "2-3", "Dusting group", "Work from top down.", True,
    ],
    [
        "Classroom Deep Clean", "Classroom", 2, "Trash Removal",
        "Empty garbage cans, clean cans inside and out, replace liners and take trash to exterior collection point.",
        5, "2", "Trash group", "Can happen while dusting starts.", True,
    ],
    [
        "Classroom Deep Clean", "Classroom", 3, "Wipe Surfaces",
        "Spot clean walls, counters, doors, handles, switches, sink area, cabinets and shelves.",
        10, "Full team", "All", "", True,
    ],
    [
        "Classroom Deep Clean", "Classroom", 4, "Desks and Chairs",
        "Wipe student desks and chairs. Move/stack furniture as needed for floor work.",
        10, "Full team", "All", "", True,
    ],
    [
        "Classroom Deep Clean", "Classroom", 5, "Windows and Sills",
        "Clean interior glass and wipe sills/trim.",
        5, "2-3", "Window group", "", True,
    ],
    [
        "Classroom Deep Clean", "Classroom", 6, "Floor Cleaning",
        "Sweep/mop hard floor areas and vacuum carpet/rug areas.",
        10, "1-2", "Floor group", "", True,
    ],
    [
        "Large Restroom Deep Clean", "Large Restroom", 1, "Spray Down",
        "Hose or spray down walls, stalls, toilets/urinals and washable surfaces.",
        5, "Full team", "All", "Use district-approved cleaner and PPE.", True,
    ],
    [
        "Large Restroom Deep Clean", "Large Restroom", 2, "Apply Cleaner/Disinfectant",
        "Apply cleaner/disinfectant and allow proper dwell time where required.",
        5, "Full team", "All", "", True,
    ],
    [
        "Large Restroom Deep Clean", "Large Restroom", 3, "Scrub/Wipe Surfaces",
        "Scrub/wipe surfaces, mirrors, partitions, walls, fixtures and high-touch areas.",
        10, "Full team", "All", "", True,
    ],
    [
        "Large Restroom Deep Clean", "Large Restroom", 4, "Machine Scrub Floor",
        "Use side-by-side scrubber or auto scrubber where appropriate.",
        10, "1-2", "Floor group", "Only where safe and appropriate for the room.", True,
    ],
    [
        "Large Restroom Deep Clean", "Large Restroom", 5, "Mop/Squeegee",
        "Mop or squeegee surfaces dry and reset room.",
        5, "Full team", "All", "", True,
    ],
    [
        "Small Restroom Deep Clean", "Small Restroom", 1, "Wipe Surfaces",
        "Wipe walls, mirror, doors, vents and high-touch areas.",
        5, "1-2", "Restroom group", "", True,
    ],
    [
        "Small Restroom Deep Clean", "Small Restroom", 2, "Scrub Toilet",
        "Scrub toilet and clean surrounding fixture area.",
        5, "1", "Restroom group", "", True,
    ],
    [
        "Small Restroom Deep Clean", "Small Restroom", 3, "Floor",
        "Sweep and mop floor. Reset supplies.",
        5, "1", "Restroom group", "", True,
    ],
    [
        "Kitchen Deep Clean", "Kitchen", 1, "Gather Supplies",
        "Gather supplies, stage equipment and coordinate team assignments.",
        5, "1", "Lead", "", True,
    ],
    [
        "Kitchen Deep Clean", "Kitchen", 2, "Mats",
        "Remove and clean mats.",
        20, "2", "Team 1", "", True,
    ],
    [
        "Kitchen Deep Clean", "Kitchen", 3, "Counters, Appliances and Walls",
        "Clean counters, appliances, walls and high-touch areas.",
        20, "2", "Team 2", "", True,
    ],
    [
        "Kitchen Deep Clean", "Kitchen", 4, "Floors",
        "Sweep and mop floors.",
        20, "2", "Team 1", "", True,
    ],
    [
        "Exterior Cleaning", "Exterior", 1, "Exterior Surfaces",
        "Remove cobwebs, dust exterior surfaces, power wash where appropriate and clean exterior windows.",
        "", "As assigned", "Exterior crew", "Adjust to site conditions.", True,
    ],
]


def _parse_print_flag(value) -> bool:
    if value is None or str(value).strip() == "":
        return True
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def load_cleaning_methods(workbook_path: str | None) -> dict[str, list[dict[str, object]]]:
    """Read printable cleaning method steps from the optional Cleaning Methods sheet."""
    if not workbook_path:
        return {}

    path = Path(workbook_path)
    if not path.exists():
        return {}

    wb = load_workbook(path, data_only=True)
    if "Cleaning Methods" not in wb.sheetnames:
        return {}

    ws = wb["Cleaning Methods"]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    index = {name: i for i, name in enumerate(headers)}

    if "Method Name" not in index:
        return {}

    methods: dict[str, list[dict[str, object]]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_method = row[index["Method Name"]] if index["Method Name"] < len(row) else ""
        method_name = str(raw_method or "").strip()
        if not method_name:
            continue

        record: dict[str, object] = {}
        for col in CLEANING_METHOD_COLUMNS:
            if col in index and index[col] < len(row):
                value = row[index[col]]
            else:
                value = ""
            record[col] = "" if value is None else value

        if not _parse_print_flag(record.get("Print on Step Sheet")):
            continue

        methods.setdefault(method_name, []).append(record)

    def sort_key(item: dict[str, object]):
        try:
            return float(str(item.get("Step Order", "") or "9999"))
        except Exception:
            return 9999

    for method_name in methods:
        methods[method_name].sort(key=sort_key)

    return dict(sorted(methods.items(), key=lambda pair: pair[0].lower()))




def ensure_cleaning_methods_sheet(workbook_path: str) -> str:
    """Create a starter Cleaning Methods sheet in an existing workbook if missing.

    If the sheet already exists, the workbook is left unchanged.
    """
    path = Path(workbook_path)
    if not path.exists():
        raise ValueError(f"Workbook not found: {path}")

    wb = load_workbook(path)
    if "Cleaning Methods" in wb.sheetnames:
        return str(path)

    ws = wb.create_sheet("Cleaning Methods")
    ws.append(CLEANING_METHOD_COLUMNS)
    _bold_header(ws)
    _freeze_header(ws)
    for row in DEFAULT_CLEANING_METHOD_ROWS:
        ws.append(row)
    _set_column_widths(ws, {
        "A": 28, "B": 18, "C": 12, "D": 28, "E": 76,
        "F": 18, "G": 18, "H": 20, "I": 48, "J": 20,
    })
    wb.save(path)
    return str(path)


def format_cleaning_method_preview(method_name: str, steps: list[dict[str, object]]) -> str:
    """Create a plain-text preview used by the UI."""
    if not steps:
        return "No printable steps found for this method."

    room_use = str(steps[0].get("Room Use", "") or "").strip()
    total_minutes = 0.0
    for step in steps:
        raw = str(step.get("Estimated Minutes", "") or "").strip()
        try:
            total_minutes += float(raw)
        except Exception:
            pass

    lines = [method_name]
    if room_use:
        lines.append(f"Room Use: {room_use}")
    if total_minutes > 0:
        lines.append(f"Estimated Step Time: {total_minutes:g} minutes")
    lines.append("")
    lines.append("Steps")

    for step in steps:
        order = str(step.get("Step Order", "") or "").strip()
        title = str(step.get("Step Title", "") or "").strip()
        detail = str(step.get("Task Detail", "") or "").strip()
        minutes = str(step.get("Estimated Minutes", "") or "").strip()
        staff = str(step.get("Staff Needed", "") or "").strip()
        role = str(step.get("Team/Role", "") or "").strip()
        notes = str(step.get("Notes", "") or "").strip()

        heading = f"{order}. {title}" if order else title
        lines.append(heading)
        if detail:
            lines.append(f"   {detail}")
        meta = []
        if minutes:
            meta.append(f"{minutes} min")
        if staff:
            meta.append(f"Staff: {staff}")
        if role:
            meta.append(f"Team/Role: {role}")
        if meta:
            lines.append("   " + " | ".join(meta))
        if notes:
            lines.append(f"   Notes: {notes}")
        lines.append("")

    return "\n".join(lines).rstrip()


def export_work_instructions_workbook(
    workbook_path: str,
    file_path: str,
    selected_method: str | None = None,
) -> str:
    """Export printable work instruction sheets from the Cleaning Methods tab."""
    methods = load_cleaning_methods(workbook_path)
    if not methods:
        raise ValueError(
            "No Cleaning Methods sheet was found, or it does not contain printable rows."
        )

    if selected_method:
        if selected_method not in methods:
            raise ValueError(f"Cleaning method not found: {selected_method}")
        methods = {selected_method: methods[selected_method]}

    target = Path(file_path)
    if target.suffix.lower() != ".xlsx":
        target = target.with_suffix(".xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for method_name, steps in methods.items():
        safe_title = "".join(ch for ch in method_name if ch not in r'[]:*?/\\')[:31] or "Instructions"
        ws = wb.create_sheet(safe_title)

        ws["A1"] = method_name
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:F1")

        room_use = str(steps[0].get("Room Use", "") or "").strip()
        total_minutes = 0.0
        for step in steps:
            try:
                total_minutes += float(str(step.get("Estimated Minutes", "") or "0"))
            except Exception:
                pass

        ws["A3"] = "Room Use"
        ws["B3"] = room_use
        ws["A4"] = "Estimated Step Time"
        ws["B4"] = f"{total_minutes:g} minutes" if total_minutes > 0 else ""

        ws.append([])
        ws.append(["Step", "Step Title", "Task Detail", "Estimated Minutes", "Staff Needed", "Team/Role", "Notes"])
        header_row = ws.max_row
        _bold_header(ws, header_row)

        for step in steps:
            ws.append([
                step.get("Step Order", ""),
                step.get("Step Title", ""),
                step.get("Task Detail", ""),
                step.get("Estimated Minutes", ""),
                step.get("Staff Needed", ""),
                step.get("Team/Role", ""),
                step.get("Notes", ""),
            ])

        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        _set_column_widths(ws, {
            "A": 10, "B": 26, "C": 70, "D": 18, "E": 18, "F": 20, "G": 42,
        })
        ws.freeze_panes = f"A{header_row + 1}"
        ws.sheet_view.showGridLines = False

    wb.save(target)
    return str(target)




def create_input_template(file_path: str = "data/Summer Scheduler Workbook.xlsx") -> str:
    """Create the current single-workbook template.

    Older versions created three separate files. The app now uses one workbook
    containing all required sheets, so this function saves exactly one workbook
    and returns that path.
    """
    target = Path(file_path)

    if target.suffix.lower() not in {".xlsx", ".xlsm"}:
        target = target / "Summer Scheduler Workbook.xlsx"

    target.parent.mkdir(parents=True, exist_ok=True)
    _create_single_summer_scheduler_template(target)
    return str(target)


def _create_single_summer_scheduler_template(path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Summer Scheduler Workbook"
    ws["A1"].font = Font(bold=True, size=14)
    instruction_rows = [
        ("A3", "What this file is for"),
        ("A4", "This one workbook contains all input data for the Summer Scheduler."),
        ("A6", "Required sheets"),
        ("A7", "Instructions, Sites, Rooms, Setup, Run Settings, Room Scope, Holidays, Cleaning Methods, Staffing and Progress."),
        ("A9", "Basic workflow"),
        ("A10", "1. Fill Sites first. Site Order controls the order schools are cleaned. 1 = first."),
        ("A11", "2. Fill Rooms with the physical room data and floor makeup."),
        ("A12", "3. Fill Room Scope for this year's plan. Blank Include fields mean ON/included. Type FALSE only for exceptions."),
        ("A13", "4. Add real non-work dates on the Holidays sheet. Saturday holidays auto-observe Friday and Sunday holidays auto-observe Monday when Observed Date is blank."),
        ("A14", "5. Update Staffing, run the scheduler and review the day-by-day output."),
        ("A15", "6. Use Cleaning Methods to maintain printable work instructions for staff."),
        ("A16", "Important rules"),
        ("A17", "Do not rename sheets or column headers. The app validates these names exactly."),
        ("A18", "Enter exact square footage when known. Fractions are allowed when estimating, such as 0.33 or 0.50."),
        ("A19", "If both square footage and fraction are entered for a floor type, square footage wins."),
        ("A20", "Scrub-Only Hard Floor covers non-strip hard flooring such as restroom tile, ceramic tile, quarry tile, sealed concrete, or VCT that should be scrubbed but not stripped/waxed."),
        ("A21", "Room Scope task toggles are exception-based: blank means ON/included. Type FALSE only for rooms or tasks you want to skip."),
        ("A22", "This applies to Include Room, Include Deep Clean, Include Strip, Include Wax, Do Carpet Cleaning and Do Exterior Cleaning."),
        ("A23", "Global settings in Run Settings turn whole task types on/off. Room Scope toggles control the room-level exceptions."),
        ("A24", "A task only runs when BOTH are true: the global setting is enabled and the room-level toggle is enabled or blank."),
        ("A25", "Do not use Room Order to skip rooms. Use Include Room = FALSE on Room Scope."),
        ("A26", "Carpet data stays in Rooms even if carpet is scheduled separately. Disable include_carpet globally when the main cleaning schedule should ignore carpet."),
        ("A27", "If editing in Google Sheets, download as Microsoft Excel (.xlsx) before loading into the app."),
        ("A29", "Holidays"),
        ("A30", "Use the Holidays sheet for actual non-work dates instead of entering a holiday count. Observed Date is optional."),
        ("A31", "If Observed Date is blank and the holiday falls on Saturday, the app observes the prior Friday. If it falls on Sunday, the app observes the following Monday."),
        ("A32", "If your district observes a different day, type that specific date in Observed Date."),
    ]
    for cell, value in instruction_rows:
        ws[cell] = value
    for cell in ["A3", "A6", "A9", "A16", "A29"]:
        ws[cell].font = Font(bold=True)
    ws.column_dimensions["A"].width = 120

    ws = wb.create_sheet("Sites")
    ws.append(["Site Name", "Site Order", "Notes"])
    _bold_header(ws)
    _freeze_header(ws)
    ws.append(["WES", 1, "Site Order controls cleaning order. 1 = first site cleaned."])
    ws.append(["JXW", 2, "Example only. Replace with your site list."])
    ws.append(["RLS", 3, "Summer school site can be placed later by using a higher Site Order."])
    _set_column_widths(ws, {"A": 18, "B": 14, "C": 72})

    ws = wb.create_sheet("Rooms")
    ws.append([
        "Site Name", "Building Name", "Zone Name", "Room Name", "Room Order",
        "Total Room SqFt", "Carpet SqFt", "Carpet Fraction", "Strip/Wax Tile SqFt",
        "Strip/Wax Tile Fraction", "Scrub-Only Hard Floor SqFt",
        "Scrub-Only Hard Floor Fraction", "Room Use", "Notes",
    ])
    _bold_header(ws)
    _freeze_header(ws)
    ws.append([
        "WES", "Main", "Classrooms", "Room 1", 1, 870, "", 0.33, "", 0.67,
        "", "", "Classroom", "Fractions must be 0 to 1. Exact sqft wins over fraction.",
    ])
    _set_column_widths(ws, {
        "A": 14, "B": 18, "C": 18, "D": 18, "E": 12, "F": 16, "G": 14,
        "H": 16, "I": 20, "J": 22, "K": 24, "L": 28, "M": 18, "N": 40,
    })

    ws = wb.create_sheet("Setup")
    ws.append(["Setting", "Value", "Description"])
    _bold_header(ws)
    _freeze_header(ws)
    setup_rows = [
        ("scheduled_shift_hours_per_day", 8.5, "Paid shift length"),
        ("lunch_hours_per_day", 0.5, "Lunch time"),
        ("break_hours_per_day", 0.5, "Total break time"),
        ("setup_hours_per_day", 0.25, "Setup/opening time"),
        ("cleanup_hours_per_day", 0.5, "Cleanup/lockup time"),
        ("productive_hours_per_staff_per_day", 6.75, "Real productive hours per worker after lunch, breaks, setup and cleanup"),
        ("deep_clean_rate_sqft_per_hour", 400, "Default deep clean production rate"),
        ("restroom_deep_clean_rate_sqft_per_hour", 250, "Deep clean rate for restrooms/bathrooms"),
        ("strip_rate_sqft_per_hour", 300, "Strip production rate"),
        ("wax_rate_sqft_per_hour", 600, "Wax production rate"),
        ("carpet_rate_sqft_per_hour", 500, "Carpet production rate"),
        ("exterior_rate_sqft_per_hour", 1000, "Exterior production rate"),
        ("wax_coats", 3, "Number of wax coats"),
        ("transition_hours_per_school", 0.0, "Inter-school logistics time"),
        ("day_start_time", "7:30 AM", "Arrival/open"),
        ("work_start_time", "7:45 AM", "Work starts"),
        ("first_break_time", "10:00 AM", "First break"),
        ("lunch_time", "12:00 PM", "Lunch"),
        ("second_break_time", "2:00 PM", "Second break"),
        ("cleanup_start_time", "3:30 PM", "Cleanup starts"),
        ("day_end_time", "4:00 PM", "End of day"),
    ]
    for row in setup_rows:
        ws.append(list(row))
    _set_column_widths(ws, {"A": 38, "B": 18, "C": 78})

    ws = wb.create_sheet("Run Settings")
    ws.append(["Setting", "Value", "Description"])
    _bold_header(ws)
    _freeze_header(ws)
    run_rows = [
        ("schedule_name", "Summer Schedule", "Name shown in app and exports"),
        ("schedule_start_date", "2026-06-15", "First workday of run. Use YYYY-MM-DD."),
        ("target_end_date", "2026-07-31", "Real calendar target date. Use YYYY-MM-DD."),
        ("target_end_day", "", "Optional override. Usually leave blank so the app calculates real workdays."),
        ("paid_holidays_in_range", 0, "Legacy fallback only. Prefer the Holidays sheet for real non-work dates."),
        ("work_on_weekends", False, "True if weekends count as workdays"),
        ("current_day", 1, "Usually 1 for a fresh run"),
        ("include_deep_clean", True, "Run deep clean tasks"),
        ("include_strip", True, "Run strip tasks"),
        ("include_wax", True, "Run wax tasks"),
        ("include_carpet", False, "Run carpet tasks"),
        ("include_exterior", False, "Run exterior tasks"),
    ]
    for row in run_rows:
        ws.append(list(row))
    _set_column_widths(ws, {"A": 30, "B": 18, "C": 80})

    ws = wb.create_sheet("Room Scope")
    ws.append([
        "Site Name", "Building Name", "Zone Name", "Room Name", "Available Day",
        "Include Room", "Include Deep Clean", "Include Strip", "Include Wax",
        "Do Carpet Cleaning", "Do Exterior Cleaning", "Notes",
    ])
    _bold_header(ws)
    _freeze_header(ws)
    ws.append(["WES", "Main", "Classrooms", "Room 1", 1, True, True, True, True, False, False, "Blank include/toggle fields default to TRUE. Type FALSE to skip."])
    _set_column_widths(ws, {
        "A": 14, "B": 18, "C": 18, "D": 18, "E": 14, "F": 14, "G": 18,
        "H": 14, "I": 14, "J": 20, "K": 20, "L": 58,
    })

    ws = wb.create_sheet("Holidays")
    ws.append(["Holiday Name", "Date", "Observed Date", "Counts As Non-Workday", "Notes"])
    _bold_header(ws)
    _freeze_header(ws)
    ws.append(["Juneteenth", "2026-06-19", "", True, "District holiday. If Observed Date is blank, Date is used unless it falls on weekend."])
    ws.append(["Independence Day", "2026-07-04", "", True, "Saturday auto-observes Friday 2026-07-03."])
    _set_column_widths(ws, {"A": 24, "B": 16, "C": 18, "D": 24, "E": 78})

    ws = wb.create_sheet("Cleaning Methods")
    ws.append(CLEANING_METHOD_COLUMNS)
    _bold_header(ws)
    _freeze_header(ws)
    for row in DEFAULT_CLEANING_METHOD_ROWS:
        ws.append(row)
    _set_column_widths(ws, {
        "A": 28, "B": 18, "C": 12, "D": 28, "E": 76,
        "F": 18, "G": 18, "H": 20, "I": 48, "J": 20,
    })

    ws = wb.create_sheet("Staffing")
    ws.append(["Day", "Available Staff", "Carpet Staff Reserved", "Absences", "Temporary Help"])
    _bold_header(ws)
    _freeze_header(ws)
    for day in range(1, 61):
        ws.append([day, 4, 0, 0, 0])
    _set_column_widths(ws, {"A": 10, "B": 18, "C": 24, "D": 12, "E": 18})

    ws = wb.create_sheet("Progress")
    ws.append(["Site Name", "Building Name", "Zone Name", "Room Name", "Task", "Hours Completed"])
    _bold_header(ws)
    _freeze_header(ws)
    _set_column_widths(ws, {"A": 14, "B": 18, "C": 18, "D": 18, "E": 20, "F": 18})

    wb.save(path)



def _create_district_facility_data_template(path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "District Facility Data"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "What this file is for"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "Use this workbook for slow-changing district facts such as sites, buildings, rooms and flooring makeup."
    ws["A6"] = "Flooring entry rules"
    ws["A6"].font = Font(bold=True)
    ws["A7"] = "Enter exact square footage when known."
    ws["A8"] = "Fractions are allowed when estimating, such as 0.33 or 0.50."
    ws["A9"] = "If both square footage and fraction are entered, square footage wins."
    ws["A10"] = "Scrub-Only VCT is its own flooring category and should not generate strip/wax work."
    ws.column_dimensions["A"].width = 110

    ws = wb.create_sheet("Sites")
    ws.append(["Site Name", "Site Order", "Notes"])
    _bold_header(ws)
    _freeze_header(ws)
    ws.append(["WES", 1, "Wright Elementary"])
    ws.append(["JXW", 2, "JX Wilson"])
    ws.append(["RLS", 3, "RL Stevens"])
    _set_column_widths(ws, {"A": 18, "B": 12, "C": 30})

    ws = wb.create_sheet("Rooms")
    ws.append(
        [
            "Site Name",
            "Building Name",
            "Zone Name",
            "Room Name",
            "Room Order",
            "Total Room SqFt",
            "Carpet SqFt",
            "Carpet Fraction",
            "Strip/Wax Tile SqFt",
            "Strip/Wax Tile Fraction",
            "Scrub-Only VCT SqFt",
            "Scrub-Only VCT Fraction",
            "Room Use",
            "Notes",
        ]
    )
    _bold_header(ws)
    _freeze_header(ws)
    ws.append(
        [
            "WES",
            "Main",
            "Classrooms",
            "Room 1",
            1,
            870,
            "",
            0.33,
            "",
            0.67,
            "",
            "",
            "Classroom",
            "",
        ]
    )
    _set_column_widths(
        ws,
        {
            "A": 14,
            "B": 18,
            "C": 18,
            "D": 18,
            "E": 12,
            "F": 16,
            "G": 14,
            "H": 16,
            "I": 18,
            "J": 20,
            "K": 18,
            "L": 20,
            "M": 16,
            "N": 24,
        },
    )
    ws.insert_rows(1)
    ws["A1"] = (
        "Enter exact square footage when known. Fractions are allowed when estimating. "
        "If both are entered, square footage wins."
    )
    ws.merge_cells("A1:N1")
    _style_instruction_row(ws, 1)
    _freeze_header(ws, "A3")

    wb.save(path)


def _create_cleaning_planning_assumptions_template(path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Cleaning Planning Assumptions"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "What this file is for"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "Use this workbook for editable planning logic such as rates, time model and wax coats."
    ws["A5"] = "Update this when real-world production data improves your assumptions."
    ws.column_dimensions["A"].width = 110

    ws = wb.create_sheet("Setup")
    ws.append(["Setting", "Value", "Description"])
    _bold_header(ws)
    _freeze_header(ws)

    rows = [
        ("scheduled_shift_hours_per_day", 8.5, "Paid shift length"),
        ("lunch_hours_per_day", 0.5, "Lunch time"),
        ("break_hours_per_day", 0.5, "Total break time"),
        ("setup_hours_per_day", 0.25, "Setup/opening time"),
        ("cleanup_hours_per_day", 0.5, "Cleanup/lockup time"),
        ("productive_hours_per_staff_per_day", 6.75, "Real productive hours per worker"),
        ("deep_clean_rate_sqft_per_hour", 400, "Deep clean production rate"),
        ("strip_rate_sqft_per_hour", 300, "Strip production rate"),
        ("wax_rate_sqft_per_hour", 600, "Wax production rate"),
        ("carpet_rate_sqft_per_hour", 500, "Carpet production rate"),
        ("exterior_rate_sqft_per_hour", 1000, "Exterior production rate"),
        ("wax_coats", 3, "Number of wax coats"),
        ("transition_hours_per_school", 0.0, "Inter-school logistics time"),
        ("day_start_time", "7:30 AM", "Arrival/open"),
        ("work_start_time", "7:45 AM", "Work starts"),
        ("first_break_time", "10:00 AM", "First break"),
        ("lunch_time", "12:00 PM", "Lunch"),
        ("second_break_time", "2:00 PM", "Second break"),
        ("cleanup_start_time", "3:30 PM", "Cleanup starts"),
        ("day_end_time", "4:00 PM", "End of day"),
    ]
    for row in rows:
        ws.append(list(row))

    _set_column_widths(ws, {"A": 34, "B": 18, "C": 42})

    wb.save(path)


def _create_summer_scheduler_run_input_template(path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Summer Scheduler Run Input"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "What this file is for"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "Use this workbook for current run / current summer data such as dates, staffing, holidays and progress."
    ws["A5"] = "This is the file you are most likely to update and rerun throughout the summer."
    ws.column_dimensions["A"].width = 110

    ws = wb.create_sheet("Run Settings")
    ws.append(["Setting", "Value", "Description"])
    _bold_header(ws)
    _freeze_header(ws)
    rows = [
        ("schedule_name", "Summer Schedule", "Name shown in app and exports"),
        ("schedule_start_date", "2026-06-01", "First workday of run"),
        ("target_end_date", "2026-06-26", "Real calendar target date"),
        ("target_end_day", "", "Optional. Leave blank to auto-calculate from target_end_date"),
        ("paid_holidays_in_range", 0, "Weekday holidays inside the run span"),
        ("work_on_weekends", False, "True if weekends count as workdays"),
        ("current_day", 1, "Usually 1 for a fresh run"),
        ("include_deep_clean", True, "Run deep clean tasks"),
        ("include_strip", True, "Run strip tasks"),
        ("include_wax", True, "Run wax tasks"),
        ("include_carpet", True, "Run carpet tasks"),
        ("include_exterior", False, "Run exterior tasks"),
    ]
    for row in rows:
        ws.append(list(row))
    _set_column_widths(ws, {"A": 30, "B": 18, "C": 46})

    ws = wb.create_sheet("Room Scope")
    ws.append(
        [
            "Site Name",
            "Building Name",
            "Zone Name",
            "Room Name",
            "Available Day",
            "Include Deep Clean",
            "Include Strip",
            "Include Wax",
            "Include Carpet",
            "Include Exterior",
            "Notes",
        ]
    )
    _bold_header(ws)
    _freeze_header(ws)
    ws.append(["WES", "Main", "Classrooms", "Room 1", 1, True, True, True, True, False, ""])
    _set_column_widths(
        ws,
        {
            "A": 14,
            "B": 18,
            "C": 18,
            "D": 18,
            "E": 14,
            "F": 18,
            "G": 14,
            "H": 14,
            "I": 16,
            "J": 16,
            "K": 24,
        },
    )

    ws = wb.create_sheet("Holidays")
    ws.append(
        [
            "Holiday Name",
            "Date",
            "Observed Date",
            "Counts As Non-Workday",
            "Notes",
        ]
    )
    _bold_header(ws)
    _freeze_header(ws)
    ws.append(["Juneteenth", "2026-06-19", "", True, "Observed date auto-calculates if blank"])
    ws.append(["Independence Day", "2026-07-04", "", True, "Saturday auto-observes Friday 2026-07-03"])
    _set_column_widths(ws, {"A": 24, "B": 16, "C": 18, "D": 24, "E": 48})

    ws = wb.create_sheet("Staffing")
    ws.append(
        [
            "Day",
            "Available Staff",
            "Carpet Staff Reserved",
            "Absences",
            "Temporary Help",
        ]
    )
    _bold_header(ws)
    _freeze_header(ws)
    for day in range(1, 41):
        ws.append([day, 4, 0, 0, 0])
    _set_column_widths(ws, {"A": 10, "B": 16, "C": 22, "D": 12, "E": 16})

    ws = wb.create_sheet("Progress")
    ws.append(
        [
            "Site Name",
            "Building Name",
            "Zone Name",
            "Room Name",
            "Task",
            "Hours Completed",
        ]
    )
    _bold_header(ws)
    _freeze_header(ws)
    _set_column_widths(ws, {"A": 14, "B": 18, "C": 18, "D": 18, "E": 20, "F": 18})

    wb.save(path)
